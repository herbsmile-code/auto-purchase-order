import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

os.makedirs('templates', exist_ok=True)
os.makedirs('output/orders', exist_ok=True)

def create_price_master():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "단가표"
    
    headers = ["품번", "품명", "규격", "매입가", "수주가", "비고"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    sample_data = [
        ["BK-101", "고급 표지 코팅지", "A4 / 250g", 3500, 5000, "무광 코팅"],
        ["BK-102", "친환경 중질지 내지", "B5 / 80g", 1200, 2000, "미색"],
        ["BK-103", "하드커버 북바인딩 키트", "양장 / 규격형", 8000, 12000, "기본 패키지"],
        ["BK-201", "맞춤형 스프링철 제본", "15mm 원형", 2500, 4000, "블랙 와이어"],
        ["BK-202", "금박/은박 후가공 세트", "표지 가공", 15000, 25000, "동판 제작 포함"],
        ["PAPER-A4", "프리미엄 복사용지", "A4 / 80g / 500매", 4500, 6500, "1박스 기준"],
        ["APP-501", "소프트커버 제작권", "표준형", 4000, 6000, "기본 수주품"],
        ["APP-502", "올컬러 도록 인쇄", "A4 / 올컬러 100p", 22000, 35000, "특수 인쇄"],
    ]
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for r_idx, row in enumerate(sample_data, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = thin_border
            if c_idx in [4, 5]: # 매입가, 수주가
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in [1, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 20
    ws.row_dimensions[1].height = 28
    
    file_path = "templates/price_master.xlsx"
    wb.save(file_path)
    print(f"Created {file_path}")

def create_order_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주서"
    
    ws.views.sheetView[0].showGridLines = True
    
    # Title
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = "발  주  서"
    title_cell.font = Font(name="맑은 고딕", size=20, bold=True, color="1E293B")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='94A3B8'),
        right=Side(style='thin', color='94A3B8'),
        top=Side(style='thin', color='94A3B8'),
        bottom=Side(style='thin', color='94A3B8')
    )
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    accent_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    table_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    table_header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    
    # Info Section
    info_layout = [
        ("A4", "발주일자", "B4", "2026-08-18"),
        ("A5", "발주번호", "B5", "PO-20260818-001"),
        ("A6", "발주처(당사)", "B6", "(주)바이브컴퍼니"),
        ("A7", "담당자/연락처", "B7", "발주팀 / 010-1234-5678"),
        ("E4", "수신(공급업체)", "F4:H4", "테스트공급사"),
        ("E5", "담당자/연락처", "F5:H5", "홍길동 / 010-9876-5432"),
        ("E6", "수령인/연락처", "F6:H6", "김수령 / 010-5555-6666"),
        ("E7", "배송지 주소", "F7:H7", "서울특별시 강남구 테헤란로 123 4층"),
    ]
    
    for item in info_layout:
        lbl_pos, lbl_text, val_pos, val_text = item
        
        lbl_cell = ws[lbl_pos]
        lbl_cell.value = lbl_text
        lbl_cell.font = Font(name="맑은 고딕", size=9, bold=True, color="334155")
        lbl_cell.fill = header_fill
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        lbl_cell.border = thin_border
        
        if ":" in val_pos:
            ws.merge_cells(val_pos)
            first_cell = ws[val_pos.split(":")[0]]
            first_cell.value = val_text
            first_cell.font = Font(name="맑은 고딕", size=9)
            first_cell.alignment = Alignment(horizontal="left", vertical="center")
            # Apply border to all merged cells
            start_col, start_row = openpyxl.utils.coordinate_to_tuple(val_pos.split(":")[0])
            end_col, end_row = openpyxl.utils.coordinate_to_tuple(val_pos.split(":")[1])
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    ws.cell(row=r, column=c).border = thin_border
        else:
            val_cell = ws[val_pos]
            val_cell.value = val_text
            val_cell.font = Font(name="맑은 고딕", size=9)
            val_cell.alignment = Alignment(horizontal="left", vertical="center")
            val_cell.border = thin_border
            
    # Table Header at Row 9
    table_headers = [
        ("A9", "No", 6),
        ("B9", "품번", 16),
        ("C9", "품명", 26),
        ("D9", "규격/옵션", 18),
        ("E9", "수량", 10),
        ("F9", "단가(매입가)", 15),
        ("G9", "공급가액", 16),
        ("H9", "비고/요청사항", 22)
    ]
    
    for pos, text, width in table_headers:
        cell = ws[pos]
        cell.value = text
        cell.fill = table_header_fill
        cell.font = table_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        col_letter = pos[0]
        ws.column_dimensions[col_letter].width = width
        
    ws.row_dimensions[9].height = 26
    
    # 10 Rows of items (Rows 10 to 19)
    for r in range(10, 20):
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=1, value=r-9).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
        
        # Qty
        c_qty = ws.cell(row=r, column=5)
        c_qty.alignment = Alignment(horizontal="right", vertical="center")
        c_qty.number_format = '#,##0'
        
        # Unit Price
        c_price = ws.cell(row=r, column=6)
        c_price.alignment = Alignment(horizontal="right", vertical="center")
        c_price.number_format = '#,##0'
        
        # Amount formula = E{r} * F{r}
        c_amt = ws.cell(row=r, column=7)
        c_amt.value = f"=IF(E{r}>0, E{r}*F{r}, 0)"
        c_amt.alignment = Alignment(horizontal="right", vertical="center")
        c_amt.number_format = '#,##0'
        
        ws.cell(row=r, column=8).alignment = Alignment(horizontal="left", vertical="center")
        
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).font = Font(name="맑은 고딕", size=9)

    # Total Row at Row 20
    ws.merge_cells("A20:F20")
    tot_lbl = ws["A20"]
    tot_lbl.value = "합  계 (VAT 별도)"
    tot_lbl.font = Font(name="맑은 고딕", size=10, bold=True, color="1E293B")
    tot_lbl.fill = accent_fill
    tot_lbl.alignment = Alignment(horizontal="center", vertical="center")
    
    for c in range(1, 7):
        ws.cell(row=20, column=c).border = thin_border
        
    tot_val = ws["G20"]
    tot_val.value = "=SUM(G10:G19)"
    tot_val.font = Font(name="맑은 고딕", size=11, bold=True, color="1E3A8A")
    tot_val.fill = accent_fill
    tot_val.alignment = Alignment(horizontal="right", vertical="center")
    tot_val.number_format = '#,##0'
    tot_val.border = thin_border
    
    ws.cell(row=20, column=8).fill = accent_fill
    ws.cell(row=20, column=8).border = thin_border
    ws.row_dimensions[20].height = 26
    
    # Memo Row at Row 22
    ws.merge_cells("A22:H24")
    memo_cell = ws["A22"]
    memo_cell.value = "【특이사항 및 전달 메모】\n* 납기일정 준수 요망\n* 파손 주의 배송 요망"
    memo_cell.font = Font(name="맑은 고딕", size=9, color="475569")
    memo_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for r in range(22, 25):
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = thin_border
            
    file_path = "templates/order_template.xlsx"
    wb.save(file_path)
    print(f"Created {file_path}")

if __name__ == "__main__":
    create_price_master()
    create_order_template()
