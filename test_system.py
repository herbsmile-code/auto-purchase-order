import os
import sys
import unittest
from fastapi.testclient import TestClient
import openpyxl

from app import (
    app,
    parse_kakao_text,
    generate_order_excel,
    append_to_erp_upload_list,
    get_price_master,
    ORDERS_DIR,
    OUTPUT_DIR
)

client = TestClient(app)

class TestKakaoOrderSystem(unittest.TestCase):

    def test_01_price_master_loading(self):
        """단가 마스터 엑셀 로딩 검증"""
        master = get_price_master()
        self.assertIn("BK-101", master)
        self.assertEqual(master["BK-101"]["buy_price"], 3500)
        self.assertEqual(master["BK-101"]["sell_price"], 5000)
        print("[PASS] 단가 마스터 로드 성공: BK-101 매입가 3500원 / 수주가 5000원")

    def test_02_kakao_parser_sample1(self):
        """카톡 샘플 1 파싱 및 단가 매칭 검증"""
        sample_text = """[발주요청]
업체: 인쇄나라
품목:
BK-101 500개
BK-202 50개
받는분: 김철수
연락처: 010-3333-7777
주소: 서울특별시 마포구 월드컵북로 120 인쇄빌딩 3층
메모: 파손주의 및 오후 2시 이전 출고 요망"""
        
        parsed = parse_kakao_text(sample_text)
        self.assertEqual(parsed["vendor_name"], "인쇄나라")
        self.assertEqual(parsed["recipient_name"], "김철수")
        self.assertEqual(parsed["recipient_phone"], "010-3333-7777")
        self.assertIn("서울특별시 마포구", parsed["recipient_address"])
        self.assertEqual(len(parsed["items"]), 2)
        
        # Item 1: BK-101 500개, 매입가 3500, 수주가 5000
        item1 = next(it for it in parsed["items"] if it["item_code"] == "BK-101")
        self.assertEqual(item1["qty"], 500)
        self.assertEqual(item1["buy_price"], 3500)
        self.assertEqual(item1["sell_price"], 5000)
        
        # Item 2: BK-202 50개, 매입가 15000, 수주가 25000
        item2 = next(it for it in parsed["items"] if it["item_code"] == "BK-202")
        self.assertEqual(item2["qty"], 50)
        self.assertEqual(item2["buy_price"], 15000)
        self.assertEqual(item2["sell_price"], 25000)
        print("[PASS] 카톡 샘플 1 파싱 및 품번/매입가/수주가 자동 매칭 성공")

    def test_03_generate_order_excel(self):
        """발주서 엑셀 생성 검증"""
        order_data = {
            "vendor_name": "인쇄나라",
            "order_no": "PO-20260818-TEST",
            "order_date": "2026-08-18",
            "recipient_name": "김철수",
            "recipient_phone": "010-3333-7777",
            "recipient_address": "서울특별시 마포구 월드컵북로 120",
            "memo": "빠른 출고 요망",
            "items": [
                {"item_code": "BK-101", "item_name": "고급 표지 코팅지", "spec": "A4 / 250g", "qty": 500, "buy_price": 3500, "sell_price": 5000, "remark": ""},
                {"item_code": "BK-202", "item_name": "금박/은박 후가공 세트", "spec": "표지 가공", "qty": 50, "buy_price": 15000, "sell_price": 25000, "remark": ""}
            ]
        }
        
        file_name = generate_order_excel(order_data)
        file_path = os.path.join(ORDERS_DIR, file_name)
        self.assertTrue(os.path.exists(file_path))
        
        # 엑셀 파일 내용 확인
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        self.assertEqual(ws["F4"].value, "인쇄나라")
        self.assertEqual(ws["B10"].value, "BK-101")
        self.assertEqual(ws["E10"].value, 500)
        self.assertEqual(ws["F10"].value, 3500)
        print(f"[PASS] 발주서 엑셀 파일 정상 생성 확인: {file_name}")

    def test_04_erp_upload_list_append(self):
        """ERP 업로드용 엑셀 행 누적 검증"""
        order_data = {
            "vendor_name": "한빛지류",
            "order_no": "PO-20260818-002",
            "order_date": "2026-08-18",
            "recipient_name": "이영희",
            "recipient_phone": "010-8888-9999",
            "recipient_address": "경기도 파주시 문발로 456",
            "memo": "거래명세서 동봉",
            "items": [
                {"item_code": "BK-102", "item_name": "친환경 중질지 내지", "spec": "B5 / 80g", "qty": 300, "buy_price": 1200, "sell_price": 2000, "remark": ""},
                {"item_code": "PAPER-A4", "item_name": "프리미엄 복사용지", "spec": "A4 / 80g", "qty": 20, "buy_price": 4500, "sell_price": 6500, "remark": ""}
            ]
        }
        
        erp_path = append_to_erp_upload_list(order_data)
        self.assertTrue(os.path.exists(erp_path))
        
        wb = openpyxl.load_workbook(erp_path, data_only=True)
        ws = wb.active
        self.assertGreaterEqual(ws.max_row, 3) # Header + at least 2 rows
        print(f"[PASS] 사내 ERP 누적 엑셀 행 정상 추가 확인 (총 행수: {ws.max_row})")

    def test_05_api_step1_flow(self):
        """1차 버튼 API 전체 연동 테스트"""
        payload = {
            "vendor_name": "동서제본",
            "order_no": "PO-20260818-003",
            "order_date": "2026-08-18",
            "recipient_name": "박민수",
            "recipient_phone": "010-5555-1234",
            "recipient_address": "부산광역시 해운대구 센텀중앙로 78",
            "memo": "도착 전 연락",
            "items": [
                {"item_code": "BK-201", "item_name": "맞춤형 스프링철 제본", "spec": "15mm 원형", "qty": 200, "buy_price": 2500, "sell_price": 4000, "remark": ""}
            ]
        }
        
        res = client.post("/api/create-order-and-erp", json=payload)
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertTrue(data["success"])
        self.assertIn("order_file", data)
        self.assertIn("mail_draft", data)
        print(f"[PASS] 1차 생성 API 성공: 생성된 발주서 = {data['order_file']}")

    def test_06_erp_summary_api(self):
        """ERP 요약 통계 API 검증"""
        res = client.get("/api/erp-summary")
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertTrue(data["exists"])
        self.assertGreater(data["total_rows"], 0)
        print(f"[PASS] ERP 요약 API 성공: 누적된 총 행 수 = {data['total_rows']}")

if __name__ == "__main__":
    unittest.main()
