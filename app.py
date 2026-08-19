import os
import json
import re
import datetime
import smtplib
from copy import copy
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import List, Optional, Dict, Any

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
ORDERS_DIR = os.path.join(OUTPUT_DIR, "orders")
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")
STATIC_DIR = os.path.join(BASE_DIR, "static")

os.makedirs(TEMPLATES_DIR, exist_ok=True)
os.makedirs(ORDERS_DIR, exist_ok=True)
os.makedirs(STATIC_DIR, exist_ok=True)

app = FastAPI(title="카카오톡 주문 원버튼 발주서 및 ERP 연동 시스템")

# ----------------- Helper Functions ----------------- #

def load_config() -> dict:
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {
        "naverworks": {
            "smtp_server": "smtp.worksmobile.com",
            "smtp_port": 465,
            "use_ssl": True,
            "sender_email": "",
            "sender_password": "",
            "sender_name": "발주담당자"
        },
        "company_info": {
            "company_name": "(주)바이브컴퍼니",
            "sender_name": "발주팀 담당자",
            "phone": "010-1234-5678",
            "email": "order@vibecompany.com"
        },
        "default_mail_template": {
            "subject": "[출고요청서] {store_name} 귀하 - {delivery_date} 출고 요청 건 ({item_summary})",
            "body": "안녕하세요, 담당자님.\n\n출고요청서 첨부하여 전달드립니다.\n확인 후 일정에 맞춰 출고 및 직배송 진행 부탁드립니다.\n\n■ 납품처: {store_name}\n■ 배송요청일: {delivery_date}\n■ 담당자: {recipient_contact}\n■ 배송장소: {recipient_address}\n\n감사합니다."
        },
        "vendors": [
            {"name": "스타리온", "email": "starion@example.com", "phone": "1588-0000"},
            {"name": "오진양행", "email": "ohjin@example.com", "phone": "02-1234-5678"}
        ]
    }

def save_config(cfg: dict):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def norm_key(s: str) -> str:
    """대소문자, 공백, 하이픈 등 특수문자를 제거하여 모델명을 유연하게 매칭하기 위한 정규화 키를 생성합니다."""
    if not s:
        return ""
    return re.sub(r'[^A-Z0-9가-힣]', '', str(s).upper())

def to_int(val: Any) -> int:
    """다양한 형식(콤마, 소수점, 통화표시 등)의 숫자 값을 정수로 안전하게 변환합니다."""
    if val is None:
        return 0
    try:
        s = str(val).replace(",", "").replace("원", "").replace("￦", "").strip()
        return int(float(s))
    except (ValueError, TypeError):
        return 0

def get_price_master() -> Dict[str, dict]:
    """price_master.xlsx 및 스타리온 판매 엑셀의 모든 시트를 자동 스캔하여
    모델명, 품번, 품명, 매입가, 할인가(=수주단가), 규격, 비고 등을 지능적으로 인덱싱합니다."""
    master = {}
    
    target_files = [
        "price_master.xlsx",
        "■스타리온&오진양행 제품 판매_2026.08.xlsx"
    ]
    
    for filename in target_files:
        file_path = os.path.join(TEMPLATES_DIR, filename)
        if not os.path.exists(file_path):
            continue
            
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                if ws.max_row is None or ws.max_row < 2:
                    continue
                    
                model_col, code_col, name_col = -1, -1, -1
                buy_col, sell_col, spec_col, remark_col = -1, -1, -1, -1
                header_row = -1
                
                # 1. 헤더 행 및 컬럼 탐색
                for r in range(1, min(20, ws.max_row + 1)):
                    for c in range(1, min(30, ws.max_column + 1)):
                        v = str(ws.cell(r, c).value or "").strip()
                        if not v:
                            continue
                        if "모델명" in v:
                            model_col = c
                            header_row = r
                        elif "품번" in v and code_col == -1:
                            code_col = c
                        elif ("품명" in v or "제품명" in v or "품목" in v) and name_col == -1:
                            name_col = c
                        elif "매입가" in v and buy_col == -1:
                            buy_col = c
                        elif ("할인가" in v or "수주" in v or "출고가" in v or "판매가" in v) and sell_col == -1:
                            sell_col = c
                        elif ("규격" in v or "사이즈" in v) and spec_col == -1:
                            spec_col = c
                        elif ("재질" in v or "색상" in v or "비고" in v) and remark_col == -1:
                            remark_col = c
                            
                    if model_col != -1 and (buy_col != -1 or sell_col != -1):
                        break
                        
                # 2. 데이터 행 파싱
                if header_row != -1 and model_col != -1:
                    for r in range(header_row + 1, ws.max_row + 1):
                        m_val = str(ws.cell(r, model_col).value or "").strip()
                        if not m_val or m_val == "None" or m_val == "0":
                            continue
                            
                        code_val = str(ws.cell(r, code_col).value or "").strip() if code_col != -1 else ""
                        name_val = str(ws.cell(r, name_col).value or "").strip() if name_col != -1 else ""
                        spec_val = str(ws.cell(r, spec_col).value or "").strip() if spec_col != -1 else ""
                        buy_p = to_int(ws.cell(r, buy_col).value) if buy_col != -1 else 0
                        sell_p = to_int(ws.cell(r, sell_col).value) if sell_col != -1 else 0
                        rem_val = str(ws.cell(r, remark_col).value or "").strip() if remark_col != -1 else ""
                        
                        if code_val == "None": code_val = ""
                        if name_val == "None": name_val = ""
                        if spec_val == "None": spec_val = ""
                        if rem_val == "None": rem_val = ""
                        
                        item_data = {
                            "model_name": m_val,
                            "item_code": code_val or m_val,
                            "item_name": name_val,
                            "spec": spec_val,
                            "buy_price": buy_p,
                            "sell_price": sell_p,
                            "remark": rem_val
                        }
                        
                        # 키별 매핑 (대문자 및 정규화 키)
                        master[m_val.upper()] = item_data
                        master[norm_key(m_val)] = item_data
                        
                        if code_val:
                            master[code_val.upper()] = item_data
                            master[norm_key(code_val)] = item_data
                            
                        if name_val:
                            master[name_val.upper()] = item_data
                            master[norm_key(name_val)] = item_data
        except Exception as e:
            print(f"[Warning] Error reading {filename}: {e}")
            
    return master

def clean_kakao_line(line: str) -> str:
    """카톡 타임스탬프, 시스템 메시지, 선행 번호/기호를 깨끗하게 제거합니다."""
    l = line.strip()
    if not l:
        return ""
    # 1. 카톡 발신자 및 시간 태그 제거: [홍길동] [오후 2:30] 또는 [오전 11:20]
    l = re.sub(r'\[[^\]]+\]\s*\[(?:오전|오후|\d{1,2}:\d{2})[^\]]*\]\s*', '', l)
    l = re.sub(r'\[(?:오전|오후|\d{1,2}:\d{2})[^\]]*\]\s*', '', l)
    # 2. 선행 번호/불릿 기호 제거: 1. 1) - * ■ ● ① ② 등 (단, 모델명이나 수량 앞 번호)
    l = re.sub(r'^(?:\d+[\.\)]|\*|-|■|●|①|②|③|④|⑤|⑥|⑦|⑧|⑨|⑩)\s*', '', l)
    return l.strip()

def is_metadata_line(line: str) -> bool:
    """해당 라인이 납품처, 배송일, 주소, 연락처 등의 메타데이터 헤더인지 판별합니다."""
    clean = clean_kakao_line(line)
    if not clean:
        return True
    meta_keywords = [
        '납품처', '납품처명', '납품일', '납품일자', '납품장소', '납품시간', '납품',
        '매장명', '매장', '지점명', '지점', '상호명', '상호', '수신처', '고객명', '업체명',
        '주소', '배송장소', '배송지', '현장주소', '현장위치', '위치',
        '현장담당', '현장담당자', '담당자', '수령인', '받는분', '연락처', '전화번호', '핸드폰', 'HP',
        '배송요청일', '배송요청', '배송일', '배송일자', '희망일', '희망시간', '출고요청일', '출고일', '도착일', '납기일',
        '특이사항', '메모', '요청사항', '비고', '전달사항',
        '모델명과 수량', '모델명 및 수량', '품목 및 수량', '품명 및 수량', '주문내역', '발주내역', '발주서', '출고요청서'
    ]
    for kw in meta_keywords:
        if clean.startswith(kw):
            after = clean[len(kw):].strip()
            if not after or after.startswith(':') or after.startswith('-') or after.startswith('：') or after.startswith('이름') or after.startswith('및'):
                return True
    if clean in ['1.', '2.', '3.', '4.', '5.', '1', '2', '3', '4', '5']:
        return True
    return False

def is_valid_item_line(raw_line: str, price_master: dict) -> bool:
    """인사말, 안부문구, 잡담, 메타데이터 등을 철저히 걸러내고 실제 장비/품목 라인만 판별합니다."""
    line = clean_kakao_line(raw_line)
    if not line:
        return False
    if line.startswith('+') or line.startswith('＋'):
        return True

    # 1. 메타데이터 라인이면 품목에서 제외
    if is_metadata_line(line):
        return False

    # 2. 대괄호 제목 라인 제외 (예: [스타리온 출고요청])
    if line.startswith('[') and line.endswith(']'):
        return False

    # 3. 인사말 및 일반 대화 문구 패턴
    greeting_patterns = [
        r'안녕(?:하세|하십|)',
        r'발주(?:하려|합니|드립|서\b| 건| 요청| 완료)',
        r'전달(?:드립|합니|해)',
        r'부탁(?:드립|합니|해|바랍)',
        r'수고(?:하세|많으|)',
        r'감사(?:합니|드립|)',
        r'확인(?:부탁|해|바랍|요망)',
        r'좋은\s*하루',
        r'오늘도\s*수고',
        r'주문(?:서\b|내역|합니|드립)',
        r'출고(?:요청|부탁|건)',
        r'^(?:사장님|대표님|부장님|과장님|대리님|담당자님|팀장님)',
        r'문의(?:드립|사항|합니)',
        r'연락(?:바랍|부탁|드립)',
    ]
    
    equipment_keywords = [
        '냉장', '냉동', '쇼케이스', '제빙기', '작업대', '선반', '씽크', '싱크',
        '테이블', '온장', '소독', '오븐', '그리들', '튀김', '믹서', '블렌더',
        '추출', '워머', '디스펜서', 'box', '박스', '음료', '아날로그', '디지털',
        '직냉', '간냉', '반찬', '토핑', '스탠드', '서브웨이', '참치', '스타리온'
    ]

    clean_line = line

    # 모델명 패턴 (영문 대문자 2자 이상 + 숫자/하이픈 조합 또는 4자 이상 영숫자)
    has_model_code = bool(re.search(r'[A-Za-z]{2,}[-_]?[A-Za-z0-9]{2,}', clean_line))
    has_equipment = any(eq in clean_line.lower() for eq in equipment_keywords)
    has_master_match = bool(price_master.get(clean_line.upper()) or price_master.get(norm_key(clean_line)))
    
    is_greeting = any(re.search(p, clean_line) for p in greeting_patterns)

    # 순수 인사말이면 탈락
    if is_greeting and not (has_equipment or has_master_match or has_model_code):
        return False

    # 장비 키워드나 모델명 또는 단가표 매칭이 있으면 통과
    if has_equipment or has_master_match or has_model_code:
        return True

    # 라인에 수량 패턴(- 1대, 2대 등)이 있고 순수 인사말이 아니면 통과
    if re.search(r'\d+\s*(?:대|개|박스|box|ea)\b', clean_line) and not is_greeting:
        return True

    return False

# ----------------- Smart Kakao Parser Logic ----------------- #

def parse_kakao_text(text: str) -> dict:
    """카카오톡 발주/출고요청 메시지를 지능적으로 파싱하여 구조화된 데이터를 생성합니다."""
    lines = [clean_kakao_line(line) for line in text.strip().split("\n") if clean_kakao_line(line)]
    price_master = get_price_master()
    config = load_config()
    
    result = {
        "vendor_name": "스타리온",
        "store_name": "",
        "delivery_date": "",
        "recipient_name": "",
        "recipient_contact": "",
        "recipient_phone": "",
        "recipient_address": "",
        "order_date": datetime.date.today().strftime("%Y-%m-%d"),
        "memo": "",
        "items": []
    }
    
    # 1. 메타데이터 파싱
    for line in lines:
        if line.startswith('[') and line.endswith(']'):
            title_content = line[1:-1].strip()
            if any(k in title_content for k in ['점', '지점', '매장', '상호', '카페', '캔', '커피']):
                if not result['store_name']:
                    result['store_name'] = title_content
            continue
            
        # 매장명 / 납품처 / 상호 / 지점
        m_store = re.search(r'(?:매장명|납품처명|납품처|매장|지점명|지점|상호명|상호|수신처|업체명)\s*[:：\-]?\s*(.+)', line)
        if m_store and not result["store_name"]:
            val = m_store.group(1).strip()
            if val and not re.match(r'^\d+\.?$', val) and not any(k in val for k in ['납품처', '모델명', '배송일', '품목', '수량']):
                result["store_name"] = val
                
        # 배송요청일 / 납품일 / 배송일
        m_date = re.search(r'(?:납품일\s*및\s*희망\s*시간|배송요청일|배송일|납품일|희망일|출고요청일|출고일|도착일|납기일|배송일자|일자)\s*[:：\-]?\s*(.+)', line)
        if m_date and not result["delivery_date"]:
            val = m_date.group(1).strip()
            if val and not re.match(r'^\d+\.?$', val):
                result["delivery_date"] = val
                
        # 주소 / 배송장소 / 배송지
        m_addr = re.search(r'(?:주소|배송장소|배송지|납품장소|현장주소|현장위치|위치)\s*[:：\-]?\s*(.+)', line)
        if m_addr and not result["recipient_address"]:
            val = m_addr.group(1).strip()
            if val and val != ":" and len(val) >= 3:
                result["recipient_address"] = val
                
        # 현장담당 / 담당자 / 전화번호
        m_contact = re.search(r'(?:현장담당\s*이름,?\s*전화번호|현장담당자?|담당자?|수령인|받는분|연락처|전화번호|핸드폰|HP)\s*[:：\-]?\s*(.+)', line)
        if m_contact and not result["recipient_contact"]:
            val = m_contact.group(1).strip()
            if val and not re.match(r'^\d+\.?$', val):
                result["recipient_contact"] = val
                ph_m = re.search(r'(01[016789]-?\d{3,4}-?\d{4})', val)
                if ph_m:
                    result["recipient_phone"] = ph_m.group(1)
                    result["recipient_name"] = val.replace(ph_m.group(1), '').strip()
                else:
                    result["recipient_name"] = val
                    
        # 특이사항 / 메모
        m_memo = re.search(r'(?:특이사항|메모|요청사항|비고|전달사항)\s*[:：\-]?\s*(.+)', line)
        if m_memo and not result["memo"]:
            val = m_memo.group(1).strip()
            if val:
                result["memo"] = val

    # 거래처 식별 (기본 스타리온, 텍스트에 오진양행 등 있을 시 반영)
    for v in config.get("vendors", []):
        if v["name"] in text:
            result["vendor_name"] = v["name"]
            break

    # fallback 주소 탐지
    if not result["recipient_address"]:
        addr_fallback = re.search(r'((?:서울|경기|인천|강원|충북|충남|대전|경북|경남|대구|전북|전남|광주|울산|부산|제주|세종)[^\n,]{6,60})', text)
        if addr_fallback:
            result["recipient_address"] = addr_fallback.group(1).strip()

    # fallback 매장명 탐색 (~점, ~지점, ~본점 등)
    if not result["store_name"]:
        for line in lines:
            if any(k in line for k in ['안녕', '발주', '감사', '주소', '010-', '배송', '모델명', '출고']):
                continue
            m_store_end = re.search(r'([가-힣A-Za-z0-9\s]{2,25}(?:점|지점|매장|호점|본점|커피|카페))\b', line)
            if m_store_end:
                result["store_name"] = m_store_end.group(1).strip()
                break

    # fallback 배송요청일 탐색
    if not result["delivery_date"]:
        date_fallback = re.search(r'(\d{1,2}\s*월\s*\d{1,2}\s*(?:[~-]\s*\d{1,2}\s*)?일|\d{1,2}\s*[/.]\s*\d{1,2}(?:\s*\(?[월화수목금토일]\)?)?)', text)
        if date_fallback:
            result["delivery_date"] = date_fallback.group(1).strip()
        else:
            result["delivery_date"] = datetime.date.today().strftime("%m월 %d일")

    # fallback 연락처 탐색
    if not result["recipient_contact"]:
        ph_fallback = re.search(r'([가-힣]{2,4}\s*)?(01[016789]-?\d{3,4}-?\d{4})', text)
        if ph_fallback:
            result["recipient_contact"] = ph_fallback.group(0).strip()
            result["recipient_phone"] = ph_fallback.group(2).strip()
            result["recipient_name"] = (ph_fallback.group(1) or '').strip()

    # 2. 품목 라인 수집 및 파싱
    item_lines = []
    in_item_sec = False
    
    for line in lines:
        if any(h in line for h in ["모델명과 수량", "모델명 및 수량", "품목 및 수량", "품명 및 수량", "2. 모델명", "2. 품목", "2. 품명", "주문내역", "발주내역"]):
            in_item_sec = True
            continue
        if in_item_sec:
            if re.match(r'^\d+\.\s*', line) and not any(k in line for k in ["모델명", "품목", "품명", "수량"]):
                in_item_sec = False
                continue
            if is_valid_item_line(line, price_master):
                item_lines.append(line)
            
    # 명시적 섹션 헤더가 없는 경우, 메타데이터가 아니고 품목 유효성을 만족하는 라인들 수집
    if not item_lines:
        for line in lines:
            if is_valid_item_line(line, price_master):
                item_lines.append(line)

    extracted_items = []
    
    for raw_line in item_lines:
        line = raw_line.strip()
        if not line:
            continue
            
        # 옵션 라인 처리 (예: '+ 1200 2단선반')
        if (line.startswith('+') or line.startswith('＋')) and extracted_items:
            extracted_items[-1]["item_name"] += f" + {line.lstrip('+＋').strip()}"
            continue

        clean_line = line

        # 1) 라인 끝의 수량 추출 (- 2대, 2대, 2개, 2 등)
        qty = 1
        m_qty = re.search(r'[-–—~:\s]+(\d+)\s*(?:대|개|박스|box|ea|EA|세트|set)?\s*$', clean_line)
        
        if m_qty:
            matched_str = m_qty.group(0)
            qty_val = int(m_qty.group(1))
            if any(u in matched_str for u in ['대', '개', '박스', 'box', 'ea', 'EA', '세트', 'set', '-', '–', '—', ':']) or re.search(r'\s+' + str(qty_val) + r'\s*$', clean_line):
                qty = qty_val
                clean_line = clean_line[:m_qty.start()].strip()

        # 2) 품명(item_name)과 모델명(model_name) 분리
        item_name = clean_line
        model_name = ""
        
        if '/' in clean_line:
            last_slash_idx = clean_line.rfind('/')
            candidate = clean_line[last_slash_idx + 1:].strip()
            # 마지막 슬래시 뒤가 모델명 패턴인지 확인
            if re.search(r'[A-Za-z0-9]', candidate):
                model_name = candidate
                item_name = clean_line[:last_slash_idx].strip()
            else:
                item_name = clean_line
        else:
            # 슬래시가 없는 경우 영문/숫자 혼합 모델명 패턴 탐색 (예: SRV12EIEVF, SR-T15B1F)
            code_match = re.search(r'([A-Za-z]{1,5}[-_]?[A-Za-z0-9\-_]{3,})', clean_line)
            if code_match:
                model_name = code_match.group(1)
                item_name = clean_line.replace(model_name, '').strip()
                item_name = re.sub(r'[\/\-:\s]+$', '', item_name).strip()

        # 모델명이 비어있고 라인 전체가 모델명 패턴인 경우
        if not model_name and re.match(r'^[A-Za-z0-9\-_]+$', clean_line):
            model_name = clean_line
            item_name = ""

        # 3) 모델명으로 단가표(price_master) 매칭 (매입가, 할인가=수주단가 조회)
        master_info = {}
        if model_name:
            master_info = price_master.get(model_name.upper(), {}) or price_master.get(norm_key(model_name), {})
        if not master_info and item_name:
            master_info = price_master.get(item_name.upper(), {}) or price_master.get(norm_key(item_name), {})
            
        buy_price = master_info.get("buy_price", 0)
        sell_price = master_info.get("sell_price", 0)
        spec = master_info.get("spec", "")
        
        # 품목명이 비어있는데 단가표에 품명이 있으면 채워줌
        if not item_name and master_info.get("item_name"):
            item_name = master_info.get("item_name")
        elif not item_name:
            item_name = model_name

        remark = "스타리온 직배송 요청"

        extracted_items.append({
            "no": len(extracted_items) + 1,
            "item_code": master_info.get("item_code") or model_name,
            "model_name": model_name,
            "item_name": item_name,
            "spec": spec,
            "qty": qty,
            "buy_price": buy_price,
            "sell_price": sell_price,
            "margin": sell_price - buy_price,
            "remark": remark
        })

    result["items"] = extracted_items
    
    # 발주번호 생성
    now = datetime.datetime.now()
    result["order_no"] = f"PO-{now.strftime('%Y%m%d')}-{now.strftime('%H%M%S')[-3:]}"
    
    return result

# ----------------- Smart Excel Generator ----------------- #

def copy_cell_style(src_cell, dst_cell):
    """원본 셀의 스타일(Font, Fill, Border, Alignment, NumberFormat)을 대상 셀에 완벽 복사합니다."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

def extract_mmdd(date_str: str) -> str:
    """배송요청일 또는 일자 텍스트에서 4자리 월일(MMDD) 포맷을 정확히 추출합니다."""
    if not date_str:
        return datetime.date.today().strftime("%m%d")
    
    # 1. 2026-08-14, 2026.08.14, 2026/08/14
    m_full = re.search(r'\d{4}[-./년\s]+(\d{1,2})[-./월\s]+(\d{1,2})', date_str)
    if m_full:
        mm = int(m_full.group(1))
        dd = int(m_full.group(2))
        return f"{mm:02d}{dd:02d}"
        
    # 2. 8월 14일, 8월 3~4일, 8월 3일
    m_kor = re.search(r'(\d{1,2})\s*월\s*(\d{1,2})', date_str)
    if m_kor:
        mm = int(m_kor.group(1))
        dd = int(m_kor.group(2))
        return f"{mm:02d}{dd:02d}"
        
    # 3. 08/14, 8/14, 08.14, 8.14
    m_slash = re.search(r'(\d{1,2})\s*[/.]\s*(\d{1,2})', date_str)
    if m_slash:
        mm = int(m_slash.group(1))
        dd = int(m_slash.group(2))
        return f"{mm:02d}{dd:02d}"
        
    # 4. 숫자 4자리 (0814)
    m_digits = re.search(r'\b(0[1-9]|1[0-2])([0-3][0-9])\b', date_str)
    if m_digits:
        return m_digits.group(0)

    # 5. fallback: 오늘 날짜
    return datetime.date.today().strftime("%m%d")

def generate_order_excel(order_data: dict) -> str:
    """사용자가 카톡에 올린 내용만 정확하게 발주서 엑셀(order_template.xlsx 서식)로 생성합니다.
    기존 템플릿에 남아있는 예시 행/잔여 데이터를 깨끗하게 클리어하고 사용자 입력 데이터만 반영합니다."""
    template_path = os.path.join(TEMPLATES_DIR, "order_template.xlsx")
    
    # 템플릿이 없으면 0814 템플릿에서 복사
    if not os.path.exists(template_path):
        sample_tpl = os.path.join(TEMPLATES_DIR, "0814_스트릿캔 구리인창점_브랜드발주서.xlsx")
        if os.path.exists(sample_tpl):
            import shutil
            shutil.copyfile(sample_tpl, template_path)
        else:
            raise FileNotFoundError("발주서 템플릿 파일이 없습니다.")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    store_name = order_data.get("store_name") or order_data.get("recipient_name") or "출고요청서"
    delivery_date = order_data.get("delivery_date") or order_data.get("order_date") or datetime.date.today().strftime("%m월 %d일")
    address = order_data.get("recipient_address") or ""
    contact = order_data.get("recipient_contact") or f"{order_data.get('recipient_name', '')} {order_data.get('recipient_phone', '')}".strip()
    memo = order_data.get("memo", "")
    items = order_data.get("items", [])

    # 1. 메타데이터 셀 스마트 탐색 및 주입
    for row in range(1, 15):
        for col in range(1, 6):
            cell_val = str(ws.cell(row=row, column=col).value or "").strip()
            
            # 납품처명 / 매장명
            if any(k in cell_val for k in ["납품처명", "납품처", "매장명"]):
                ws.cell(row=row, column=col + 1, value=store_name)
                
            # 배송요청일 / 납품일
            elif any(k in cell_val for k in ["배송요청일", "납품일", "희망일"]):
                ws.cell(row=row, column=col + 1, value=delivery_date)
                
            # 배송장소 / 주소
            elif any(k in cell_val for k in ["배송장소", "주소", "배송지"]):
                ws.cell(row=row, column=col + 1, value=address)
                
            # 담당자 / 현장담당 / 연락처
            elif any(k in cell_val for k in ["담당자", "현장담당", "연락처"]):
                ws.cell(row=row, column=col + 1, value=contact)

    # 2. 품목 테이블 헤더 행 및 컬럼 탐색
    header_row = None
    col_map = {}
    
    for r in range(10, 20):
        row_vals = [str(ws.cell(row=r, column=c).value or "").strip().upper() for c in range(1, 10)]
        if any(h in row_vals for h in ["NO", "품목", "품명", "모델명", "품번"]):
            header_row = r
            for c in range(1, 10):
                c_val = str(ws.cell(row=r, column=c).value or "").strip().upper()
                if "NO" in c_val:
                    col_map["no"] = c
                elif "품목" in c_val or "품명" in c_val:
                    col_map["item_name"] = c
                elif "모델명" in c_val or "품번" in c_val:
                    col_map["model_name"] = c
                elif "수량" in c_val:
                    col_map["qty"] = c
                elif "비고" in c_val:
                    col_map["remark"] = c
                elif "단가" in c_val or "매입" in c_val:
                    col_map["price"] = c
            break

    # 기본 컬럼 매핑 보정 (A=NO, B=품목, C=모델명, D=수량, E=비고)
    if not header_row:
        header_row = 14
        col_map = {"no": 1, "item_name": 2, "model_name": 3, "qty": 4, "remark": 5}

    data_start_row = header_row + 1

    # 템플릿의 첫 번째 데이터 행 서식을 스타일 참조로 저장
    ref_styles = {}
    max_c = max(col_map.values(), default=5)
    for col_idx in range(1, max_c + 1):
        ref_styles[col_idx] = ws.cell(row=data_start_row, column=col_idx)

    # 기존 템플릿의 데이터 영역(예시 행들)을 깨끗하게 클리어
    max_existing_row = max(ws.max_row, data_start_row + 10)
    for r in range(data_start_row, max_existing_row + 1):
        # '특이사항' 헤더 행은 보존
        c1_val = str(ws.cell(row=r, column=1).value or "").strip()
        if "특이사항" in c1_val or "메모" in c1_val:
            ws.cell(row=r, column=1, value="")
            ws.cell(row=r, column=3, value="")
            continue
        for c in range(1, max_c + 2):
            ws.cell(row=r, column=c, value=None)

    # 사용자가 올린 품목 데이터만 정확하게 주입
    for idx, item in enumerate(items):
        current_r = data_start_row + idx
        ws.row_dimensions[current_r].height = 24

        # NO
        if "no" in col_map:
            c_no = ws.cell(row=current_r, column=col_map["no"], value=str(idx + 1))
            copy_cell_style(ref_styles.get(col_map["no"], c_no), c_no)
            c_no.alignment = Alignment(horizontal="center", vertical="center")

        # 품목
        if "item_name" in col_map:
            c_name = ws.cell(row=current_r, column=col_map["item_name"], value=item.get("item_name", ""))
            copy_cell_style(ref_styles.get(col_map["item_name"], c_name), c_name)
            c_name.alignment = Alignment(horizontal="center", vertical="center")

        # 모델명
        if "model_name" in col_map:
            c_model = ws.cell(row=current_r, column=col_map["model_name"], value=item.get("model_name") or item.get("item_code", ""))
            copy_cell_style(ref_styles.get(col_map["model_name"], c_model), c_model)
            c_model.alignment = Alignment(horizontal="center", vertical="center")

        # 수량
        if "qty" in col_map:
            c_qty = ws.cell(row=current_r, column=col_map["qty"], value=int(item.get("qty", 1)))
            copy_cell_style(ref_styles.get(col_map["qty"], c_qty), c_qty)
            c_qty.alignment = Alignment(horizontal="center", vertical="center")
            c_qty.number_format = '#,##0'

        # 비고
        if "remark" in col_map:
            c_remark = ws.cell(row=current_r, column=col_map["remark"], value=item.get("remark", "스타리온 직배송 요청"))
            copy_cell_style(ref_styles.get(col_map["remark"], c_remark), c_remark)
            c_remark.alignment = Alignment(horizontal="center", vertical="center")

    # 3. 특이사항 메모 기입 (데이터 테이블 바로 아래 배치)
    memo_row = data_start_row + len(items) + 1
    c_memo_label = ws.cell(row=memo_row, column=1, value="특이사항")
    c_memo_label.font = Font(name="맑은 고딕", size=10, bold=True)
    c_memo_label.alignment = Alignment(horizontal="center", vertical="center")
    if memo:
        ws.cell(row=memo_row, column=3, value=memo)

    # 파일 저장 (요청 형식: {MMDD}_{납품처(매장명)}_브랜드발주서.xlsx)
    mmdd = extract_mmdd(delivery_date or order_data.get("order_date") or "")
    safe_name = re.sub(r'[\/:*?"<>|]', '_', store_name).strip() if store_name else "미지정"
    file_name = f"{mmdd}_{safe_name}_브랜드발주서.xlsx"
    out_file_path = os.path.join(ORDERS_DIR, file_name)
    
    wb.save(out_file_path)
    return file_name

def get_pil_font(size=14, bold=False):
    """윈도우 기본 폰트(맑은 고딕)를 로드합니다."""
    font_names = ['malgunbd.ttf' if bold else 'malgun.ttf', 'arialbd.ttf' if bold else 'arial.ttf']
    font_dirs = ['C:/Windows/Fonts', 'C:/WINNT/Fonts']
    for fd in font_dirs:
        for fn in font_names:
            fp = os.path.join(fd, fn)
            if os.path.exists(fp):
                try:
                    return ImageFont.truetype(fp, size)
                except Exception:
                    pass
    return ImageFont.load_default()

def generate_order_image(order_data: dict) -> str:
    """출고요청서 정보를 100% 동일한 비주얼의 고해상도 PNG 이미지로 렌더링하여 저장합니다."""
    store_name = order_data.get("store_name") or order_data.get("recipient_name") or "출고요청서"
    delivery_date = order_data.get("delivery_date") or order_data.get("order_date") or datetime.date.today().strftime("%m월 %d일")
    address = order_data.get("recipient_address") or ""
    contact = order_data.get("recipient_contact") or f"{order_data.get('recipient_name', '')} {order_data.get('recipient_phone', '')}".strip()
    memo = order_data.get("memo", "")
    items = order_data.get("items", [])

    width = 900
    bold_font = get_pil_font(14, bold=True)
    normal_font = get_pil_font(14, bold=False)

    num_items = max(len(items), 1)
    row_height = 36
    table_height = (num_items + 1) * row_height
    height = 200 + table_height + 80

    img = Image.new('RGB', (width, height), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)

    # 1. 상단 메타데이터
    x_start = 20
    y = 20
    draw.text((x_start, y), '납품처명:    ', fill=(0, 0, 0), font=bold_font)
    draw.text((x_start + 110, y), store_name, fill=(0, 0, 0), font=normal_font)

    y += 28
    draw.text((x_start, y), '배송요청일 : ', fill=(0, 0, 0), font=bold_font)
    draw.text((x_start + 110, y), delivery_date, fill=(0, 0, 220), font=bold_font)

    y += 28
    draw.text((x_start, y), '배송장소 :   ', fill=(0, 0, 0), font=bold_font)
    draw.text((x_start + 110, y), address, fill=(0, 0, 0), font=normal_font)

    y += 28
    draw.text((x_start, y), '담당자 :     ', fill=(0, 0, 0), font=bold_font)
    draw.text((x_start + 110, y), contact, fill=(0, 0, 0), font=normal_font)

    y += 45
    draw.text((x_start, y), '아래와 같이 기기류에 대해 견적하오니 검토 바랍니다.', fill=(0, 0, 0), font=bold_font)

    # 2. 품목 테이블
    y += 35
    table_x = 20
    table_w = width - 40
    col_widths = [60, 360, 170, 70, 200]
    cols_x = [table_x]
    for w in col_widths:
        cols_x.append(cols_x[-1] + w)

    # 헤더
    draw.rectangle([table_x, y, table_x + table_w, y + row_height], fill=(166, 166, 166), outline=(100, 100, 100))
    headers = ['NO', '품목', '모델명', '수량', '비고']
    for i, h in enumerate(headers):
        cx = cols_x[i] + col_widths[i] // 2
        cy = y + row_height // 2
        draw.text((cx, cy), h, fill=(0, 0, 0), font=bold_font, anchor='mm')
        draw.line([cols_x[i], y, cols_x[i], y + row_height], fill=(100, 100, 100), width=1)
    draw.line([cols_x[-1], y, cols_x[-1], y + row_height], fill=(100, 100, 100), width=1)

    # 행들
    curr_y = y + row_height
    for idx, it in enumerate(items, 1):
        draw.rectangle([table_x, curr_y, table_x + table_w, curr_y + row_height], fill=(255, 255, 255), outline=(150, 150, 150))
        # NO
        draw.text((cols_x[0] + col_widths[0] // 2, curr_y + row_height // 2), str(idx), fill=(0, 0, 0), font=normal_font, anchor='mm')
        # 품목
        draw.text((cols_x[1] + 10, curr_y + row_height // 2), it.get('item_name', ''), fill=(0, 0, 0), font=normal_font, anchor='lm')
        # 모델명
        m_name = it.get('model_name') or it.get('item_code', '')
        draw.text((cols_x[2] + col_widths[2] // 2, curr_y + row_height // 2), m_name, fill=(0, 0, 0), font=normal_font, anchor='mm')
        # 수량
        draw.text((cols_x[3] + col_widths[3] // 2, curr_y + row_height // 2), str(it.get('qty', 1)), fill=(0, 0, 0), font=normal_font, anchor='mm')
        
        for cx in cols_x:
            draw.line([cx, curr_y, cx, curr_y + row_height], fill=(150, 150, 150), width=1)
        curr_y += row_height

    # 비고 병합 박스
    if len(items) > 0:
        remark_text = items[0].get('remark') or '스타리온 직배송 요청'
        remark_cy = y + row_height + (len(items) * row_height) // 2
        draw.rectangle([cols_x[4], y + row_height, cols_x[5], curr_y], fill=(255, 255, 255), outline=(150, 150, 150))
        draw.text((cols_x[4] + col_widths[4] // 2, remark_cy), remark_text, fill=(0, 0, 0), font=normal_font, anchor='mm')

    # 3. 특이사항 Box
    curr_y += 15
    box_h = 45
    draw.rectangle([table_x, curr_y, table_x + 300, curr_y + box_h], fill=(255, 255, 255), outline=(100, 100, 100))
    draw.rectangle([table_x + 300, curr_y, table_x + table_w, curr_y + box_h], fill=(255, 255, 255), outline=(100, 100, 100))
    draw.text((table_x + 150, curr_y + box_h // 2), '특이사항', fill=(0, 0, 0), font=bold_font, anchor='mm')
    draw.text((table_x + 315, curr_y + box_h // 2), memo, fill=(0, 0, 0), font=normal_font, anchor='lm')

    mmdd = extract_mmdd(delivery_date)
    clean_store = re.sub(r'[\\/:*?"<>|]', '', store_name)
    image_filename = f"{mmdd}_{clean_store}_출고요청서.png"
    image_path = os.path.join(ORDERS_DIR, image_filename)
    
    os.makedirs(ORDERS_DIR, exist_ok=True)
    img.save(image_path)
    return image_filename

def extract_sheet_tab_name(delivery_date: str = "", order_date: str = "") -> str:
    """배송요청일 또는 발주일자에서 엑셀 시트 탭 이름(예: 26.8, 26.9, 26.10)을 추출합니다."""
    now = datetime.date.today()
    y_str = str(now.year)[-2:] # '26'
    
    # 1. 배송요청일 우선 분석 (예: 8월 14일, 9월 5일, 8.24, 2026-09-10)
    if delivery_date:
        m_full = re.search(r'(\d{2,4})[-./년\s]+(\d{1,2})[-./월\s]+(\d{1,2})', delivery_date)
        if m_full:
            y = m_full.group(1)
            y_2d = y[-2:] if len(y) >= 2 else y_str
            return f"{y_2d}.{int(m_full.group(2))}"
            
        m_kor = re.search(r'(\d{1,2})\s*월', delivery_date)
        if m_kor:
            return f"{y_str}.{int(m_kor.group(1))}"
            
        m_dot = re.search(r'(\d{1,2})\s*[/.]\s*(\d{1,2})', delivery_date)
        if m_dot:
            m = int(m_dot.group(1))
            if 1 <= m <= 12:
                return f"{y_str}.{m}"

    # 2. 발주일자 분석 (예: 2026-08-19)
    if order_date:
        m_full = re.search(r'(\d{2,4})[-./년\s]+(\d{1,2})', order_date)
        if m_full:
            y = m_full.group(1)
            y_2d = y[-2:] if len(y) >= 2 else y_str
            return f"{y_2d}.{int(m_full.group(2))}"
            
    return f"{y_str}.{now.month}"

def format_erp_sheet_header(ws, headers, col_widths):
    """ERP 시트의 16개 컬럼 헤더 및 너비를 사용자 서식에 맞게 완벽하게 스타일링합니다."""
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num, value=headers[col_num - 1])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

def append_to_erp_upload_list(order_data: dict, replace_all: bool = False) -> str:
    """사내 ERP 업로드용 엑셀 파일(erp_upload_list.xlsx)에 월별 탭(26.8, 26.9 등)을 자동 분리하여 기록합니다.
    - 서식: 사용자 16개 컬럼 서식 100% 보존
    - 거래처, 거래처명: 빈 공란(None)으로 유지
    - 배송요청일이 9월이면 '26.9' 시트 탭을 자동 생성하여 9월 건만 기록"""
    erp_file_path = os.path.join(OUTPUT_DIR, "erp_upload_list.xlsx")
    
    headers = [
        "발주일자", "배송요청일", "거래처", "거래처명", "납품처(매장명)",
        "품명", "모델명", "수량", "매입단가", "매입합계",
        "수주단가", "수주합계", "예상마진", "배송방법", "현장담당자", "배송장소"
    ]
    
    col_widths = {
        'A': 13.0, 'B': 15.0, 'C': 8.38, 'D': 13.0, 'E': 20.0,
        'F': 35.0, 'G': 18.0, 'H': 8.0,  'I': 13.0, 'J': 14.0,
        'K': 13.0, 'L': 14.0, 'M': 14.0, 'N': 20.0, 'O': 20.0, 'P': 35.0
    }
    
    delivery_date = order_data.get("delivery_date", "")
    order_date = order_data.get("order_date", datetime.date.today().strftime("%Y-%m-%d"))
    tab_name = extract_sheet_tab_name(delivery_date, order_date)
    
    if replace_all:
        # 단독 생성 시 워크북을 새로 만들어 해당 월 탭만 단독 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tab_name
        format_erp_sheet_header(ws, headers, col_widths)
    else:
        # 가장 최신의 ERP 파일 로드 (최신 백업본 우선, 그 후 원본)
        backup_files = sorted(
            [os.path.join(OUTPUT_DIR, f) for f in os.listdir(OUTPUT_DIR) if f.startswith("erp_upload_list_") and f.endswith(".xlsx")],
            key=os.path.getmtime,
            reverse=True
        )
        candidate_paths = backup_files + [erp_file_path]
        
        wb = None
        for p in candidate_paths:
            if os.path.exists(p):
                try:
                    wb = openpyxl.load_workbook(p)
                    break
                except Exception:
                    continue
        if wb is None:
            wb = openpyxl.Workbook()
            
        # 기존 단일 디폴트 시트(Sheet 또는 출고_ERP_누적리스트) 마이그레이션
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            ws = wb["Sheet"]
            ws.title = tab_name
            format_erp_sheet_header(ws, headers, col_widths)
        elif "출고_ERP_누적리스트" in wb.sheetnames:
            if tab_name not in wb.sheetnames:
                ws = wb["출고_ERP_누적리스트"]
                ws.title = tab_name
                format_erp_sheet_header(ws, headers, col_widths)
            else:
                del wb["출고_ERP_누적리스트"]
            
        # 대상 월별 탭(시트) 가져오기 또는 새로 생성
        if tab_name in wb.sheetnames:
            ws = wb[tab_name]
        else:
            ws = wb.create_sheet(title=tab_name)
            format_erp_sheet_header(ws, headers, col_widths)
            
    # 헤더가 없거나 1행이 비어있는 경우 헤더 서식 주입
    if ws.cell(1, 1).value != "발주일자":
        format_erp_sheet_header(ws, headers, col_widths)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    store_name = order_data.get("store_name") or order_data.get("recipient_name", "")
    customer_name = order_data.get("customer_name") or None
    contact = order_data.get("recipient_contact") or f"{order_data.get('recipient_name', '')} {order_data.get('recipient_phone', '')}".strip()
    address = order_data.get("recipient_address", "")
    memo = order_data.get("memo", "")
    
    for item in order_data.get("items", []):
        qty = int(item.get("qty", 1))
        buy_p = int(item.get("buy_price", 0))
        sell_p = int(item.get("sell_price", 0))
        buy_total = qty * buy_p
        sell_total = qty * sell_p
        margin = sell_total - buy_total
        delivery_method = item.get("remark") or "스타리온 직배송 요청"
        
        # 16개 컬럼 데이터 (거래처 코드는 빈칸 None, 거래처명은 사용자 입력값 customer_name)
        row_data = [
            order_date,          # A: 발주일자
            delivery_date,       # B: 배송요청일
            None,                # C: 거래처 (빈공란)
            customer_name,       # D: 거래처명 (사용자가 수기 입력한 값, 없으면 None)
            store_name,          # E: 납품처(매장명)
            item.get("item_name", ""),                             # F: 품명
            item.get("model_name") or item.get("item_code", ""),   # G: 모델명
            qty,                 # H: 수량
            buy_p,               # I: 매입단가
            buy_total,           # J: 매입합계
            sell_p,              # K: 수주단가
            sell_total,          # L: 수주합계
            margin,              # M: 예상마진
            delivery_method,     # N: 배송방법
            contact,             # O: 현장담당자
            address              # P: 배송장소
        ]
        
        ws.append(row_data)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 20
        
        for c_idx in range(1, len(headers) + 1):
            c = ws.cell(row=current_row, column=c_idx)
            c.font = Font(name="맑은 고딕", size=10)
            c.border = thin_border
            
            # 숫자 및 정렬 서식
            if c_idx in [8, 9, 10, 11, 12]: # 수량, 매입단가, 매입합계, 수주단가, 수주합계
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx == 13: # 예상마진
                c.number_format = '[Blue][>=1]+#,##0;[Red][<0]-#,##0;0'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in [1, 2, 3, 4, 7, 14]: # 발주일자, 배송요청일, 거래처, 거래처명, 모델명, 배송방법
                c.alignment = Alignment(horizontal="center", vertical="center")
            else: # 납품처, 품명, 현장담당자, 배송장소
                c.alignment = Alignment(horizontal="left", vertical="center")
                
    format_erp_sheet_header(ws, headers, col_widths)
        
    saved_path = erp_file_path
    try:
        wb.save(erp_file_path)
    except PermissionError:
        alt_path = os.path.join(OUTPUT_DIR, f"erp_upload_list_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx")
        wb.save(alt_path)
        saved_path = alt_path
        print(f"[Warning] erp_upload_list.xlsx is currently open. Saved to {alt_path} instead.")
    return saved_path

# ----------------- Naver Works Mail Sender ----------------- #

def send_naverworks_email(to_email: str, subject: str, body_text: str, attachment_filename: str) -> dict:
    """네이버웍스 SMTP(smtp.worksmobile.com:465)를 사용하여 발주서 첨부 메일을 발송합니다.
    쉼표/세미콜론으로 구분된 다중 수신자를 완벽하게 지원합니다."""
    config = load_config()
    nw = config.get("naverworks", {})
    
    smtp_server = nw.get("smtp_server", "smtp.worksmobile.com")
    smtp_port = nw.get("smtp_port", 465)
    sender_email = nw.get("sender_email", "").strip()
    sender_password = nw.get("sender_password", "").strip()
    sender_name = nw.get("sender_name", "김진영")
    
    if not sender_email or not sender_password:
        raise ValueError("네이버웍스 이메일 계정 정보(이메일, 비밀번호)가 설정되어 있지 않습니다. 설정 화면에서 입력해주세요.")
        
    # 다중 수신자 이메일 목록 파싱
    to_list = [e.strip() for e in re.split(r'[,;\s]+', str(to_email).strip()) if e.strip() and '@' in e]
    if not to_list:
        raise ValueError("수신자 이메일 주소가 올바르지 않거나 비어 있습니다.")
        
    msg = MIMEMultipart()
    msg["From"] = f"{sender_name} <{sender_email}>"
    msg["To"] = ", ".join(to_list)
    msg["Subject"] = subject
    msg["Date"] = datetime.datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0900")
    
    msg.attach(MIMEText(body_text, "plain", "utf-8"))
    
    attach_path = os.path.join(ORDERS_DIR, attachment_filename)
    if os.path.exists(attach_path):
        with open(attach_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename*=UTF-8''{attachment_filename}"
            )
            msg.attach(part)
    else:
        raise FileNotFoundError(f"첨부할 발주서 파일이 존재하지 않습니다: {attachment_filename}")
        
    try:
        server = smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=15)
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, to_list, msg.as_string())
        server.quit()
        return {"success": True, "message": f"'{', '.join(to_list)}'(으)로 발주서 메일이 성공적으로 전송되었습니다."}
    except Exception as e:
        raise RuntimeError(f"네이버웍스 메일 발송 실패: {str(e)}")

# ----------------- API Endpoints ----------------- #

class ParseRequest(BaseModel):
    text: str

class OrderItem(BaseModel):
    no: Optional[int] = 1
    item_code: Optional[str] = ""
    model_name: Optional[str] = ""
    item_name: Optional[str] = ""
    spec: Optional[str] = ""
    qty: int = 1
    buy_price: Optional[int] = 0
    sell_price: Optional[int] = 0
    remark: Optional[str] = "스타리온 직배송 요청"

class CreateOrderRequest(BaseModel):
    vendor_name: Optional[str] = "스타리온"
    customer_name: Optional[str] = ""    # 수기 입력 거래처명 (ERP 업로드용)
    store_name: Optional[str] = ""
    delivery_date: Optional[str] = ""
    order_no: Optional[str] = None
    order_date: Optional[str] = None
    recipient_name: Optional[str] = ""
    recipient_contact: Optional[str] = ""
    recipient_phone: Optional[str] = ""
    recipient_address: Optional[str] = ""
    memo: Optional[str] = ""
    items: List[OrderItem]
    replace_erp: Optional[bool] = False

class SendMailRequest(BaseModel):
    to_email: str
    subject: str
    body: str
    attachment_filename: str

class TestMailRequest(BaseModel):
    to_email: str

@app.post("/api/parse-order")
async def api_parse_order(req: ParseRequest):
    """카카오톡 주문 텍스트를 파싱하여 납품처, 배송일, 모델명, 품목, 수량 및 단가(매입가/수주단가)를 구조화합니다."""
    if not req.text.strip():
        raise HTTPException(status_code=400, detail="카톡 주문 내용을 입력해주세요.")
    data = parse_kakao_text(req.text)
    
    # 거래처 기본 이메일 매칭 (기본 3명 수신자)
    config = load_config()
    default_recip = config.get("default_mail_template", {}).get("default_recipient", "hj.seo@starion.co.kr, gscheon@starion.co.kr, cth@ohjin.co.kr")
    data["vendor_email"] = default_recip
    data["customer_name"] = ""
    
    return {"success": True, "data": data}

import string

class SafeDict(dict):
    """템플릿 포맷팅 시 누락된 키가 있어도 KeyError를 발생시키지 않고 원본 {key}를 유지합니다."""
    def __missing__(self, key):
        return "{" + key + "}"

def safe_format(template_str: str, **kwargs) -> str:
    """안전하게 문자열 포맷팅을 수행합니다."""
    if not template_str:
        return ""
    try:
        return string.Formatter().vformat(template_str, (), SafeDict(**kwargs))
    except Exception:
        return template_str

@app.post("/api/create-order-and-erp")
async def api_create_order_and_erp(order: CreateOrderRequest):
    """[1차 버튼] 서식을 100% 보존하여 출고요청서 엑셀 및 고화질 이미지를 생성하고 사내 ERP 엑셀에 반영합니다."""
    order_dict = order.model_dump() if hasattr(order, 'model_dump') else order.dict()

    # 1. 거래처 발송용 출고요청서 엑셀 생성 (카톡에 올린 내용만 입력)
    excel_filename = generate_order_excel(order_dict)
    
    # 2. 메일 본문 삽입용 출고요청서 고화질 이미지 생성
    image_filename = generate_order_image(order_dict)
    
    # 3. 사내 ERP 누적 엑셀 행 추가 (replace_erp가 True면 현재 주문 건으로 대체)
    append_to_erp_upload_list(order_dict, replace_all=order.replace_erp)
    
    # 4. 메일 템플릿 기본 내용 생성 (스크린샷 포맷 100% 적용)
    config = load_config()
    mail_tpl = config.get("default_mail_template", {})
    comp_info = config.get("company_info", {})
    
    items = order_dict.get("items", [])
    first_item = items[0]["item_name"] if items else "물품"
    item_summary = f"{first_item} 외 {len(items)-1}건" if len(items) > 1 else first_item
    
    to_email = mail_tpl.get("default_recipient") or "hj.seo@starion.co.kr, gscheon@starion.co.kr, cth@ohjin.co.kr"
            
    store_name = order_dict.get("store_name") or order_dict.get("recipient_name") or "매장"
    delivery_date = order_dict.get("delivery_date") or order_dict.get("order_date") or datetime.date.today().strftime("%m월 %d일")
    order_date = order_dict.get("order_date") or datetime.date.today().strftime("%Y-%m-%d")
    recipient_name = order_dict.get("recipient_name") or ""
    recipient_phone = order_dict.get("recipient_phone") or ""
    contact = order_dict.get("recipient_contact") or f"{recipient_name} {recipient_phone}".strip()
    address = order_dict.get("recipient_address") or ""
    vendor_name = order_dict.get("vendor_name", "스타리온")
    department = comp_info.get("department", "브랜드영업팀")
    sender_name = comp_info.get("sender_name", "김진영")

    mmdd = extract_mmdd(delivery_date)
    # 단축 배송일 (예: 0824 -> 8/24)
    m_short = re.search(r'(\d{1,2})\s*월?\s*(\d{1,2})', delivery_date)
    if m_short:
        delivery_date_short = f"{int(m_short.group(1))}/{int(m_short.group(2))}"
    elif len(mmdd) == 4:
        delivery_date_short = f"{int(mmdd[:2])}/{int(mmdd[2:])}"
    else:
        delivery_date_short = delivery_date

    # 1. 텍스트 품목 테이블 생성 (스크린샷 1:1)
    item_table_lines = []
    item_table_lines.append(f"{'NO':<4} | {'품목':<45} | {'모델명':<18} | {'수량':<6} | {'비고'}")
    item_table_lines.append("-" * 95)
    for idx, it in enumerate(items, 1):
        i_name = it.get("item_name", "")
        m_name = it.get("model_name") or it.get("item_code", "")
        q = f"{it.get('qty', 1)}"
        rm = it.get("remark") or "스타리온 직배송 요청"
        item_table_lines.append(f"{idx:<4} | {i_name:<45} | {m_name:<18} | {q:<6} | {rm}")
    item_table_lines.append("-" * 95)
    item_table_text = "\n".join(item_table_lines)

    memo_text = order_dict.get("memo") or ""

    format_vars = {
        "mmdd": mmdd,
        "store_name": store_name,
        "date": order_date,
        "order_date": order_date,
        "delivery_date": delivery_date,
        "delivery_date_short": delivery_date_short,
        "item_summary": item_summary,
        "item_table_text": item_table_text,
        "recipient_name": recipient_name or store_name,
        "recipient_phone": recipient_phone or contact,
        "recipient_contact": contact,
        "recipient_address": address,
        "vendor_name": vendor_name,
        "department": department,
        "sender_name": sender_name,
        "memo": memo_text
    }

    # 기본 본문 템플릿
    raw_subject = mail_tpl.get("subject") or "{mmdd}_출고요청_브랜드팀 발주요청 드립니다. ({store_name})"
    raw_body = mail_tpl.get("body") or (
        "안녕하세요, 현지님.\n"
        "브랜드영업팀 김진영입니다.\n\n"
        "{delivery_date_short} 아래와 같이 스타리온 직배송 요청 드립니다.\n"
        "확인 후 진행 부탁드리겠습니다.\n\n"
    )

    subject = safe_format(raw_subject, **format_vars)
    body = safe_format(raw_body, **format_vars)

    # 2. 네이버웍스 웹메일 리치에디터용 HTML 테이블 서식 (스크린샷 비주얼 100% 재현)
    html_rows = []
    num_items = len(items)
    remark_default = items[0].get("remark") if items else "스타리온 직배송 요청"
    if not remark_default:
        remark_default = "스타리온 직배송 요청"

    for idx, it in enumerate(items, 1):
        i_name = it.get("item_name", "")
        m_name = it.get("model_name") or it.get("item_code", "")
        q = it.get("qty", 1)
        
        # 첫 번째 행에서 비고 rowspan 병합
        remark_td = ""
        if idx == 1:
            remark_td = f'<td rowspan="{num_items}" style="border: 1px solid #777; padding: 8px 10px; text-align: center; vertical-align: middle; background-color: #fff;">{remark_default}</td>'
        
        html_rows.append(f"""
        <tr>
          <td style="border: 1px solid #777; padding: 6px; text-align: center; background-color: #fff;">{idx}</td>
          <td style="border: 1px solid #777; padding: 6px 12px; text-align: left; background-color: #fff;">{i_name}</td>
          <td style="border: 1px solid #777; padding: 6px; text-align: center; background-color: #fff;">{m_name}</td>
          <td style="border: 1px solid #777; padding: 6px; text-align: center; background-color: #fff;">{q}</td>
          {remark_td}
        </tr>
        """)

    body_html = f"""
    <div style="font-family: '맑은 고딕', 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif; font-size: 14px; color: #111; line-height: 1.7;">
      <div style="margin-bottom: 20px;">
        <div><strong>납품처명:</strong>&nbsp;&nbsp;&nbsp;&nbsp;{store_name}</div>
        <div><strong>배송요청일 :</strong>&nbsp;<span style="color: #0033cc; font-weight: bold;">{delivery_date}</span></div>
        <div><strong>배송장소 :</strong>&nbsp;&nbsp;&nbsp;{address}</div>
        <div><strong>담당자 :</strong>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{contact}</div>
      </div>
      
      <div style="margin-top: 25px; margin-bottom: 12px; font-weight: bold;">
        아래와 같이 기기류에 대해 견적하오니 검토 바랍니다.
      </div>
      
      <table style="border-collapse: collapse; width: 100%; max-width: 820px; border: 1px solid #777; font-size: 13.5px;">
        <thead>
          <tr style="background-color: #a6a6a6; color: #000; font-weight: bold; height: 32px;">
            <th style="border: 1px solid #777; padding: 6px; width: 45px; text-align: center;">NO</th>
            <th style="border: 1px solid #777; padding: 6px 10px; text-align: center;">품목</th>
            <th style="border: 1px solid #777; padding: 6px; width: 150px; text-align: center;">모델명</th>
            <th style="border: 1px solid #777; padding: 6px; width: 60px; text-align: center;">수량</th>
            <th style="border: 1px solid #777; padding: 6px; width: 180px; text-align: center;">비고</th>
          </tr>
        </thead>
        <tbody>
          {"".join(html_rows)}
        </tbody>
      </table>
      
      <table style="border-collapse: collapse; width: 100%; max-width: 820px; border: 1px solid #777; border-top: none; font-size: 13.5px; margin-top: 10px;">
        <tr>
          <td style="border: 1px solid #777; width: 220px; text-align: center; font-weight: bold; padding: 10px; background-color: #fff;">특이사항</td>
          <td style="border: 1px solid #777; padding: 10px; background-color: #fff;">{memo_text}</td>
        </tr>
      </table>
    </div>
    """
    
    return {
        "success": True,
        "message": "출고요청서 엑셀 및 사내 ERP 리스트 생성이 완료되었습니다.",
        "order_file": excel_filename,
        "mail_draft": {
            "to_email": to_email,
            "subject": subject,
            "body": body,
            "body_html": body_html,
            "attachment_filename": excel_filename
        }
    }

@app.post("/api/clear-erp-list")
async def api_clear_erp_list():
    """사내 ERP 누적 엑셀(erp_upload_list.xlsx)의 모든 월별 탭의 데이터 행을 초기화합니다."""
    erp_file_path = os.path.join(OUTPUT_DIR, "erp_upload_list.xlsx")
    headers = [
        "발주일자", "배송요청일", "거래처", "거래처명", "납품처(매장명)",
        "품명", "모델명", "수량", "매입단가", "매입합계",
        "수주단가", "수주합계", "예상마진", "배송방법", "현장담당자", "배송장소"
    ]
    col_widths = {
        'A': 13.0, 'B': 15.0, 'C': 8.38, 'D': 13.0, 'E': 20.0,
        'F': 35.0, 'G': 18.0, 'H': 8.0,  'I': 13.0, 'J': 14.0,
        'K': 13.0, 'L': 14.0, 'M': 14.0, 'N': 20.0, 'O': 20.0, 'P': 35.0
    }
    
    now = datetime.date.today()
    default_tab = f"{str(now.year)[-2:]}.{now.month}"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = default_tab
    format_erp_sheet_header(ws, headers, col_widths)
    
    wb.save(erp_file_path)
    return {"success": True, "message": f"ERP 누적 리스트가 초기화되었습니다. (기본 탭: {default_tab})"}

@app.post("/api/send-mail")
async def api_send_mail(req: SendMailRequest):
    """[2차 버튼] 네이버웍스를 통해 발주서 첨부 메일을 발송합니다."""
    try:
        result = send_naverworks_email(
            to_email=req.to_email,
            subject=req.subject,
            body_text=req.body,
            attachment_filename=req.attachment_filename
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/test-naverworks")
async def api_test_naverworks(req: TestMailRequest):
    """네이버웍스 SMTP 연동 테스트 메일을 발송합니다."""
    try:
        config = load_config()
        nw = config.get("naverworks", {})
        sender_email = nw.get("sender_email", "")
        if not sender_email or not nw.get("sender_password"):
            raise ValueError("네이버웍스 계정 정보(이메일, 비밀번호)가 설정되어 있지 않습니다.")
            
        msg = MIMEMultipart()
        msg["From"] = f"{nw.get('sender_name', '김진영')} <{sender_email}>"
        msg["To"] = req.to_email
        msg["Subject"] = "[테스트] 네이버웍스 연동 테스트 메일입니다."
        msg.attach(MIMEText("네이버웍스 SMTP 연동이 정상적으로 작동하고 있습니다.\n\n발주 자동화 시스템 준비 완료!", "plain", "utf-8"))
        
        server = smtplib.SMTP_SSL(nw.get("smtp_server", "smtp.worksmobile.com"), nw.get("smtp_port", 465), timeout=10)
        server.login(sender_email, nw.get("sender_password"))
        server.sendmail(sender_email, [req.to_email], msg.as_string())
        server.quit()
        return {"success": True, "message": f"'{req.to_email}'(으)로 테스트 메일이 성공적으로 전송되었습니다!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"네이버웍스 연결 실패: {str(e)}")

# ----------------- Download & Explorer Endpoints ----------------- #

@app.post("/api/open-order-folder")
async def api_open_order_folder(filename: Optional[str] = None):
    """윈도우 탐색기에서 발주서 폴더를 열고, 파일명이 있으면 해당 파일을 바로 선택(하이라이트)합니다."""
    try:
        import subprocess
        os.makedirs(ORDERS_DIR, exist_ok=True)
        if filename:
            target_file = os.path.join(ORDERS_DIR, filename)
            if os.path.exists(target_file):
                subprocess.Popen(f'explorer /select,"{target_file}"')
                return {"success": True, "message": f"'{filename}' 파일을 탐색기에서 선택했습니다."}
        subprocess.Popen(f'explorer "{ORDERS_DIR}"')
        return {"success": True, "message": "발주서 폴더를 열었습니다."}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/send-mail")
async def api_send_mail(req: SendMailRequest):
    """[2차 버튼] 네이버웍스를 통해 발주서 첨부 메일을 발송합니다."""
    try:
        result = send_naverworks_email(
            to_email=req.to_email,
            subject=req.subject,
            body_text=req.body,
            attachment_filename=req.attachment_filename
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/test-naverworks")
async def api_test_naverworks(req: TestMailRequest):
    """네이버웍스 SMTP 연동 테스트 메일을 발송합니다."""
    try:
        config = load_config()
        nw = config.get("naverworks", {})
        sender_email = nw.get("sender_email", "")
        if not sender_email or not nw.get("sender_password"):
            raise ValueError("네이버웍스 계정 정보(이메일, 비밀번호)가 설정되어 있지 않습니다.")
            
        msg = MIMEMultipart()
        msg["From"] = f"{nw.get('sender_name', '김진영')} <{sender_email}>"
        msg["To"] = req.to_email
        msg["Subject"] = "[테스트] 네이버웍스 연동 테스트 메일입니다."
        msg.attach(MIMEText("네이버웍스 SMTP 연동이 정상적으로 작동하고 있습니다.\n\n발주 자동화 시스템 준비 완료!", "plain", "utf-8"))
        
        server = smtplib.SMTP_SSL(nw.get("smtp_server", "smtp.worksmobile.com"), nw.get("smtp_port", 465), timeout=10)
        server.login(sender_email, nw.get("sender_password"))
        server.sendmail(sender_email, [req.to_email], msg.as_string())
        server.quit()
        return {"success": True, "message": f"'{req.to_email}'(으)로 테스트 메일이 성공적으로 전송되었습니다!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"네이버웍스 연결 실패: {str(e)}")

# ----------------- Download & Explorer Endpoints ----------------- #

@app.post("/api/open-order-folder")
async def api_open_order_folder(filename: Optional[str] = None):
    """윈도우 탐색기에서 발주서 폴더를 열고, 파일명이 있으면 해당 파일을 바로 선택(하이라이트)합니다."""
    try:
        import subprocess
        os.makedirs(ORDERS_DIR, exist_ok=True)
        if filename:
            target_file = os.path.join(ORDERS_DIR, filename)
            if os.path.exists(target_file):
                subprocess.Popen(f'explorer /select,"{target_file}"')
                return {"success": True, "message": f"'{filename}' 파일을 탐색기에서 선택했습니다."}
        subprocess.Popen(f'explorer "{ORDERS_DIR}"')
        return {"success": True, "message": "발주서 폴더를 열었습니다."}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.get("/api/download/order/{filename}")
async def download_order_file(filename: str):
    file_path = os.path.join(ORDERS_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")
    return FileResponse(file_path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/download/image/{filename}")
async def download_order_image_file(filename: str):
    file_path = os.path.join(ORDERS_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="이미지 파일을 찾을 수 없습니다.")
    return FileResponse(file_path, filename=filename, media_type="image/png")

@app.get("/api/download/erp-list")
async def download_erp_list():
    file_path = os.path.join(OUTPUT_DIR, "erp_upload_list.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="ERP 누적 엑셀 파일이 아직 없습니다. 발주서를 먼저 생성해주세요.")
    return FileResponse(file_path, filename="erp_upload_list.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/download/price-master")
async def download_price_master():
    file_path = os.path.join(TEMPLATES_DIR, "price_master.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="단가표 파일이 없습니다.")
    return FileResponse(file_path, filename="price_master.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/download/order-template")
async def download_order_template():
    file_path = os.path.join(TEMPLATES_DIR, "order_template.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="발주서 템플릿 파일이 없습니다.")
    return FileResponse(file_path, filename="order_template.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ----------------- Master & Config Management ----------------- #

@app.get("/api/config")
async def api_get_config():
    return load_config()

@app.post("/api/config")
async def api_save_config(config_data: dict):
    save_config(config_data)
    return {"success": True, "message": "설정이 저장되었습니다."}

@app.get("/api/price-master")
async def api_get_price_master():
    master = get_price_master()
    unique_items = {}
    for k, v in master.items():
        code = v["item_code"]
        if code not in unique_items:
            unique_items[code] = v
    return {"success": True, "items": list(unique_items.values())}

@app.post("/api/upload/price-master")
async def api_upload_price_master(file: UploadFile = File(...)):
    file_path = os.path.join(TEMPLATES_DIR, "price_master.xlsx")
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    return {"success": True, "message": "단가표(price_master.xlsx)가 성공적으로 업데이트되었습니다."}

@app.post("/api/upload/order-template")
async def api_upload_order_template(file: UploadFile = File(...)):
    file_path = os.path.join(TEMPLATES_DIR, "order_template.xlsx")
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    return {"success": True, "message": "발주서 양식(order_template.xlsx)이 성공적으로 업데이트되었습니다."}

@app.get("/api/erp-summary")
async def api_get_erp_summary():
    erp_file_path = os.path.join(OUTPUT_DIR, "erp_upload_list.xlsx")
    if not os.path.exists(erp_file_path):
        return {"exists": False, "total_rows": 0, "recent_items": []}
        
    wb = openpyxl.load_workbook(erp_file_path, data_only=True)
    all_rows = []
    
    # 모든 월별 탭(시트) 순회
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(min_row=2, values_only=True))
        for r in sheet_rows:
            if r and any(r):
                all_rows.append((sheet_name, r))
    
    recent = []
    for sheet_name, r in reversed(all_rows[-15:]):
        # 16개 컬럼 구조:
        # 0: 발주일자, 1: 배송요청일, 2: 거래처, 3: 거래처명, 4: 납품처,
        # 5: 품명, 6: 모델명, 7: 수량, 8: 매입단가, 9: 매입합계,
        # 10: 수주단가, 11: 수주합계, 12: 예상마진, 13: 배송방법, 14: 현장담당자, 15: 배송장소
        if len(r) >= 13:
            recent.append({
                "tab": sheet_name,
                "date": str(r[0] or ''),
                "order_no": f"[{sheet_name}]",
                "store": str(r[4] or ''),
                "item_name": str(r[5] or ''),
                "model_name": str(r[6] or ''),
                "qty": r[7] if len(r) > 7 else 1,
                "buy_price": r[8] if len(r) > 8 else 0,
                "sell_price": r[10] if len(r) > 10 else 0,
                "margin": r[12] if len(r) > 12 else 0
            })
            
    return {
        "exists": True,
        "total_rows": len(all_rows),
        "recent_items": recent
    }

# ----------------- Static Files ----------------- #
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.get("/")
async def serve_index():
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))

if __name__ == "__main__":
    import uvicorn
    print("Server started at http://127.0.0.1:8000")
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
